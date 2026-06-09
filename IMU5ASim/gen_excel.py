"""Tạo file Excel mô tả chức năng phần mềm IMU 5A Simulator."""

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Chức năng phần mềm"

# ── Màu sắc ──────────────────────────────────────────────────────────────────
COL_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # xanh đậm
COL_GROUP_FILL    = PatternFill("solid", fgColor="2E75B6")   # xanh nhạt
COL_ODD_FILL      = PatternFill("solid", fgColor="DEEAF1")   # xanh rất nhạt
COL_EVEN_FILL     = PatternFill("solid", fgColor="FFFFFF")   # trắng
COL_TITLE_FILL    = PatternFill("solid", fgColor="1F4E79")

THIN  = Side(style="thin",   color="4472C4")
MED   = Side(style="medium", color="1F4E79")
thin_border  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
group_border = Border(left=MED,  right=MED,  top=MED,  bottom=THIN)

WHITE_BOLD  = Font(name="Times New Roman", bold=True, color="FFFFFF", size=13)
WHITE_NORM  = Font(name="Times New Roman", bold=False, color="FFFFFF", size=11)
DARK_BOLD   = Font(name="Times New Roman", bold=True, color="1F4E79", size=11)
DARK_NORM   = Font(name="Times New Roman", color="1A1A1A", size=11)
DARK_GROUP  = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11)

CENTER_MID  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_MID    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Độ rộng cột ──────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 65

# ── Dòng tiêu đề tổng ────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 30
ws.row_dimensions[3].height = 18

ws.merge_cells("A1:C1")
c = ws["A1"]
c.value = "VIỆN HÀNG KHÔNG VŨ TRỤ VIETTEL"
c.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11, italic=True)
c.fill = COL_TITLE_FILL
c.alignment = CENTER_MID

ws.merge_cells("A2:C2")
c = ws["A2"]
c.value = "MÔ TẢ CHỨC NĂNG PHẦN MỀM MÔ PHỎNG IMU 5A SIMULATOR"
c.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=14)
c.fill = COL_TITLE_FILL
c.alignment = CENTER_MID

ws.merge_cells("A3:C3")
c = ws["A3"]
c.value = "Giao thức: RS422 | Tần số: 125 Hz | Packet: 42 byte big-endian"
c.font = Font(name="Times New Roman", italic=True, color="BDD7EE", size=10)
c.fill = COL_TITLE_FILL
c.alignment = CENTER_MID

# ── Header cột ───────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 22
for col, val in zip("ABC", ["STT", "Tên nút / Nhóm chức năng", "Chức năng"]):
    c = ws.cell(row=4, column={"A":1,"B":2,"C":3}[col], value=val)
    c.font   = WHITE_BOLD
    c.fill   = COL_HEADER_FILL
    c.alignment = CENTER_MID
    c.border = thin_border

# ── Dữ liệu ──────────────────────────────────────────────────────────────────
# Format: (stt, ten_nhom, chuc_nang, is_group)
# is_group=True → dòng tiêu đề nhóm (span cả B+C, màu xanh đậm)

ROWS = [
    # ── Nhóm 1 ──
    (None, "NHÓM 1: NHẬP DỮ LIỆU (Data Input)", None, True),
    (1,  "Chọn file CSV (Browse)",
         "Mở hộp thoại chọn file CSV chứa dữ liệu IMU. "
         "File phải có 12 cột theo thứ tự: Index, Roll(°), Pitch(°), Yaw(°), "
         "Gx(dps), Gy(dps), Gz(dps), Ax(g), Ay(g), Az(g), Temp(°C), USW. "
         "Sau khi load thành công, hiển thị số frame và thời gian tương ứng (giây).",
         False),
    (2,  "Tạo file mẫu (Template)",
         "Tạo và lưu file CSV mẫu 125 dòng (tương đương 1 giây dữ liệu ở 125 Hz). "
         "File mẫu dùng để mở trong Excel, chỉnh sửa giá trị rồi nạp lại bằng Browse.",
         False),
    (3,  "Số frame đã nạp",
         "Nhãn hiển thị số lượng frame (dòng dữ liệu) đã đọc thành công từ file CSV "
         "và thời gian mô phỏng tương ứng. Ví dụ: \"Loaded: 250 frames (2.00 s)\".",
         False),
    (4,  "Chế độ phát — Loop / One-shot",
         "Chọn chế độ phát dữ liệu:\n"
         "• Loop: Khi hết dữ liệu, quay lại frame đầu và phát liên tục.\n"
         "• One-shot: Phát hết file một lần rồi tự động dừng.",
         False),

    # ── Nhóm 2 ──
    (None, "NHÓM 2: KẾT NỐI RS422 (Transport)", None, True),
    (5,  "Cổng serial (Port)",
         "Danh sách thả xuống chọn cổng COM kết nối thiết bị USB-RS422. "
         "Tự động quét và hiển thị tất cả cổng đang sẵn có trên hệ thống.",
         False),
    (6,  "Làm mới cổng (⟳)",
         "Quét lại danh sách cổng COM đang hoạt động trên máy tính. "
         "Nhấn sau khi cắm hoặc rút thiết bị USB-RS422 để cập nhật danh sách.",
         False),
    (7,  "Tốc độ truyền (Baud rate)",
         "Chọn tốc độ baud của cổng serial RS422. "
         "Các giá trị hỗ trợ: 921600 (mặc định), 460800, 230400, 115200, 57600 baud. "
         "Phải khớp với cấu hình bên nhận.",
         False),
    (8,  "Hệ số tỉ lệ góc (Angle LSB)",
         "Hệ số chuyển đổi từ giá trị vật lý sang giá trị nguyên trong packet. "
         "Mặc định: 0.01 deg/LSB → raw = round(góc / 0.01). "
         "Lưu dưới dạng int16 big-endian tại offset [12–17].",
         False),
    (9,  "Hệ số tỉ lệ Gyro (Gyro LSB)",
         "Hệ số chuyển đổi vận tốc góc. "
         "Mặc định: 0.001 dps/LSB → raw = round(dps / 0.001). "
         "Lưu dưới dạng int32 big-endian tại offset [18–29].",
         False),
    (10, "Hệ số tỉ lệ Accel (Accel LSB)",
         "Hệ số chuyển đổi gia tốc. "
         "Mặc định: 0.001 g/LSB → raw = round(g / 0.001). "
         "Lưu dưới dạng int16 big-endian tại offset [30–35].",
         False),
    (11, "Hệ số tỉ lệ nhiệt độ (Temp LSB)",
         "Hệ số chuyển đổi nhiệt độ. "
         "Mặc định: 0.01 °C/LSB → raw = round(°C / 0.01). "
         "Lưu dưới dạng int16 big-endian tại offset [36–37].",
         False),

    # ── Nhóm 3 ──
    (None, "NHÓM 3: ĐIỀU KHIỂN & TRẠNG THÁI (Control & Status)", None, True),
    (12, "Nút Start (▶)",
         "Bắt đầu quá trình mô phỏng: mở cổng serial với cấu hình đã chọn, "
         "khởi động luồng QThread ưu tiên cao nhất (TimeCriticalPriority), "
         "gửi packet 42 byte ở tần số 125 Hz bằng busy-wait spin-lock. "
         "Yêu cầu đã chọn file CSV hợp lệ và cổng COM hợp lệ.",
         False),
    (13, "Nút Stop (■)",
         "Dừng quá trình mô phỏng: gửi tín hiệu dừng an toàn đến luồng "
         "(không kill đột ngột), đóng cổng serial sau khi luồng kết thúc.",
         False),
    (14, "Nút Guide (?)",
         "Mở hộp thoại hướng dẫn sử dụng gồm 7 tab. "
         "Có thể mở nhanh bằng phím tắt F1 hoặc menu Help > Guide.",
         False),
    (15, "Thanh tiến trình (Progress bar)",
         "Hiển thị phần trăm tiến trình phát dữ liệu trong file CSV "
         "(0–100%). Cập nhật mỗi 12 frame (~96 ms). "
         "Hữu ích nhất ở chế độ One-shot để biết còn bao lâu hết dữ liệu.",
         False),
    (16, "Hàng dữ liệu hiện tại (Row)",
         "Hiển thị chỉ số dòng CSV đang phát và tổng số dòng. "
         "Ví dụ: \"Row: 87 / 250\". Cập nhật mỗi 12 frame.",
         False),
    (17, "Tổng packet đã gửi (Sent)",
         "Đếm tổng số packet 42 byte đã gửi thành công qua cổng serial "
         "kể từ lần nhấn Start gần nhất. Bao gồm cả các vòng lặp (Loop mode).",
         False),
    (18, "Tần số thực tế (Rate)",
         "Hiển thị tần số gửi packet thực đo được (Hz) dựa trên số packet gửi "
         "chia thời gian đã chạy. Mục tiêu: 125.0 Hz, jitter < 0.5 ms.",
         False),

    # ── Nhóm 4 ──
    (None, "NHÓM 4: BIỂU ĐỒ THỜI GIAN THỰC (Signal Preview)", None, True),
    (19, "Biểu đồ Orientation (góc định hướng)",
         "Đồ thị cuộn hiển thị 5 giây gần nhất (50 điểm, cập nhật ~10 Hz):\n"
         "• Roll (đỏ): góc lăn, đơn vị độ\n"
         "• Pitch (xanh lá): góc ngẩng, đơn vị độ\n"
         "• Yaw (xanh dương): góc hướng, đơn vị độ",
         False),
    (20, "Biểu đồ Gyro (vận tốc góc)",
         "Đồ thị cuộn hiển thị 5 giây gần nhất:\n"
         "• Gx (đỏ): vận tốc góc trục X, đơn vị dps (degree per second)\n"
         "• Gy (xanh lá): vận tốc góc trục Y\n"
         "• Gz (xanh dương): vận tốc góc trục Z",
         False),
    (21, "Biểu đồ Accel (gia tốc)",
         "Đồ thị cuộn hiển thị 5 giây gần nhất:\n"
         "• Ax (đỏ): gia tốc trục X, đơn vị g (9.81 m/s²)\n"
         "• Ay (xanh lá): gia tốc trục Y\n"
         "• Az (xanh dương): gia tốc trục Z",
         False),
    (22, "Biểu đồ Temperature (nhiệt độ)",
         "Đồ thị cuộn hiển thị nhiệt độ bên trong cảm biến IMU "
         "trong 5 giây gần nhất, đơn vị °C. Màu cam.",
         False),

    # ── Nhóm 5 ──
    (None, "NHÓM 5: PACKET HEX MONITOR (Giám sát packet)", None, True),
    (23, "Packet Hex Monitor (log)",
         "Hiển thị hex dump của packet 42 byte gần nhất được gửi, kèm giá trị vật lý "
         "đã giải mã (Roll, Pitch, Yaw, Gx, Gy, Gz, Ax, Ay, Az, Temp). "
         "Cập nhật ~5 Hz. Giúp xác minh packet có đúng định dạng không. "
         "Log tự xóa khi vượt 500 dòng để tránh tràn bộ nhớ.",
         False),
    (24, "Nút Clear log",
         "Xóa toàn bộ nội dung vùng hiển thị hex dump. "
         "Log sẽ tiếp tục cập nhật sau khi xóa.",
         False),

    # ── Nhóm 6 ──
    (None, "NHÓM 6: MENU VÀ HƯỚNG DẪN (Help)", None, True),
    (25, "Help > Guide (F1) — Tab Quick Start",
         "Hướng dẫn 4 bước khởi động nhanh: "
         "(1) Chọn file CSV → (2) Chọn cổng COM và baud rate → "
         "(3) Nhấn Start → (4) Quan sát biểu đồ và log.",
         False),
    (26, "Help > Guide — Tab CSV Template",
         "Mô tả format file CSV đầu vào: tên cột, đơn vị, cách export từ Excel, "
         "cách dùng nút Template để tạo file mẫu.",
         False),
    (27, "Help > Guide — Tab Port & Baud Setup",
         "Hướng dẫn tìm cổng COM, cài driver USB-RS422, "
         "chọn baud rate phù hợp với thiết bị nhận.",
         False),
    (28, "Help > Guide — Tab Scale Factors",
         "Giải thích ý nghĩa của LSB (Least Significant Bit), "
         "công thức tính giá trị nguyên từ giá trị vật lý, "
         "khi nào cần thay đổi hệ số tỉ lệ.",
         False),
    (29, "Help > Guide — Tab Signal Preview",
         "Hướng dẫn đọc các biểu đồ thời gian thực: "
         "cửa sổ cuộn 5 giây, màu sắc các kênh, tự động căn chỉnh trục Y.",
         False),
    (30, "Help > Guide — Tab Packet Hex Monitor",
         "Hướng dẫn đọc hex dump 42 byte: vị trí từng trường dữ liệu, "
         "cách xác minh checksum thủ công (sum bytes[2..39] mod 2¹⁶).",
         False),
    (31, "Help > Guide — Tab Timing & Jitter",
         "Giải thích cơ chế timing 125 Hz bằng busy-wait QueryPerformanceCounter, "
         "mục tiêu jitter < 0.5 ms, cách tối ưu khi CPU load cao "
         "(đóng ứng dụng nền, không dùng máy ảo).",
         False),
]

# ── Ghi các dòng dữ liệu ─────────────────────────────────────────────────────
row = 5
odd = True  # xen kẽ màu

for stt, ten, chuc_nang, is_group in ROWS:
    ws.row_dimensions[row].height = 15 if is_group else None  # auto height nếu không phải group

    if is_group:
        # Dòng tiêu đề nhóm: gộp B+C, màu xanh đậm
        ws.merge_cells(f"B{row}:C{row}")
        ca = ws.cell(row=row, column=1, value="")
        ca.fill = COL_GROUP_FILL
        ca.border = thin_border

        cb = ws.cell(row=row, column=2, value=ten)
        cb.font      = DARK_GROUP
        cb.fill      = COL_GROUP_FILL
        cb.alignment = LEFT_MID
        cb.border    = thin_border

        ws.row_dimensions[row].height = 20
        odd = True  # reset xen kẽ sau mỗi nhóm
    else:
        fill = COL_ODD_FILL if odd else COL_EVEN_FILL

        ca = ws.cell(row=row, column=1, value=stt)
        ca.font      = DARK_BOLD
        ca.fill      = fill
        ca.alignment = CENTER_MID
        ca.border    = thin_border

        cb = ws.cell(row=row, column=2, value=ten)
        cb.font      = DARK_BOLD
        cb.fill      = fill
        cb.alignment = LEFT_MID
        cb.border    = thin_border

        cc = ws.cell(row=row, column=3, value=chuc_nang)
        cc.font      = DARK_NORM
        cc.fill      = fill
        cc.alignment = LEFT_MID
        cc.border    = thin_border

        odd = not odd

    row += 1

# ── Footer ────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 18
ws.merge_cells(f"A{row}:C{row}")
cf = ws.cell(row=row, column=1,
             value="Packet 42 byte RS422: [AA][55][01][95][Len×2][NumPkg=05][DataList×5]"
                   "[Roll×2][Pitch×2][Yaw×2][Gx×4][Gy×4][Gz×4][Ax×2][Ay×2][Az×2][Temp×2][USW×2][Chksum×2]")
cf.font      = Font(name="Courier New", size=9, color="1F4E79", italic=True)
cf.fill      = PatternFill("solid", fgColor="EBF3FB")
cf.alignment = LEFT_MID
cf.border    = Border(left=MED, right=MED, top=THIN, bottom=MED)

# ── Freeze header ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ── Lưu file ─────────────────────────────────────────────────────────────────
out = r"d:\Code\SimulateSignal\IMU5ASim\Chuc_nang_IMU5ASim.xlsx"
wb.save(out)
print(f"Saved: {out}")
